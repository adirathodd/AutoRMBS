import pdfplumber
import json
import openai
import os
from dotenv import load_dotenv
import re
from openpyxl import load_workbook
import sys

load_dotenv()

# List of fields to extract
FIELDS = (
    "Closing Date, First Payment Date, Day Count System, Payment Frequency, "
    "Payment Frequency Add. Description, Description, Rate Adjustment Frequency, Initial Asset Balance, "
    "Current Prepaid Balance, Asset Amortization Type, WA Fixed Rate, Prepayment Type, "
    "Fixed Prepayment Rate, Default Rate, Recoverable, Original Term, Loss Multiple, Base Losses, "
    "Remaining Term, Discount Rate, WA Original Amortization Term, WA Original Balloon Payment Month, "
    "WA Original Interest Only Period, WA Original Interest Capitalization Period, WALA, Recoveries Lag"
)

CHUNK_SIZE = 3000  # Adjust chunk size (in characters) to avoid going over model token limits

def extract_text_from_pdf(pdf_path):
    """Extract text from a PDF file."""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
    return text

def chunk_text(text, chunk_size=CHUNK_SIZE):
    """
    Break the text into smaller chunks to handle large inputs.
    Adjust chunk_size as needed based on your token/model constraints.
    """
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start = end
    return chunks

def parse_chunk_with_openai(chunk, api_key, azure_endpoint, deployment_id, api_version):
    """Send a chunk of text to Azure OpenAI for parsing key-value pairs."""
    openai.api_type = "azure"
    openai.api_key = api_key
    openai.api_base = azure_endpoint
    openai.api_version = api_version

    system_prompt = (
        "You are a data extraction AI. You must return only valid JSON containing "
        "any of the following fields if found: " + FIELDS + ". If a field is not found, omit it. "
        "Do not include any extra keys, text, or commentary. Output only valid JSON."
    )

    user_prompt = (
        f"Text to parse:\n{chunk}\n\n"
        f"Fields to extract: {FIELDS}\n\n"
        "Return only valid JSON with only the fields you find."
    )

    response = openai.ChatCompletion.create(
        engine=deployment_id,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        temperature=0.5,
        max_tokens=1200,
    )

    structured_data = response["choices"][0]["message"]["content"]
    # Attempt to parse JSON from the chunk's response
    try:
        chunk_json = json.loads(structured_data)
    except json.JSONDecodeError:
        return {"error": "Failed to parse JSON", "response": structured_data}
    return chunk_json

def merge_dicts(main_dict, new_dict):
    """
    Merge key-value pairs from new_dict into main_dict.
    If a key already exists, decide how to handle collisions (overwrite or combine).
    """
    for key, value in new_dict.items():
        # If the field doesn't exist or it's an error, overwrite
        if key not in main_dict or key == "error":
            main_dict[key] = value
        else:
            # If both are strings and different, we can combine them
            if isinstance(main_dict[key], str) and isinstance(value, str) and main_dict[key] != value:
                main_dict[key] += f"; {value}"
            else:
                main_dict[key] = value
    return main_dict

def pdf_to_json(pdf_path, api_key, azure_endpoint, deployment_id, api_version):
    """
    Convert a large PDF file into structured JSON data by chunking the extracted text.
    """
    extracted_text = extract_text_from_pdf(pdf_path)
    chunks = chunk_text(extracted_text, CHUNK_SIZE)

    final_data = {}
    for i, chunk in enumerate(chunks):
        chunk_data = parse_chunk_with_openai(
            chunk, api_key, azure_endpoint, deployment_id, api_version
        )
        if "error" not in chunk_data:
            final_data = merge_dicts(final_data, chunk_data)
        else:
            pass
    return final_data

# Regex Patterns
dollar_pattern = re.compile(r'^\$\s*[\d,]+(\.\d+)?$')   # e.g. "$ 550,462,191", "$123,456.78"
percent_pattern = re.compile(r'^[\d,]+(\.\d+)?%$')      # e.g. "12%", "12.34%", "1,234.56%"
numeric_pattern = re.compile(r'^[\d,]+(\.\d+)?$')       # e.g. "1234", "1,234.56", "1234.56"

def save_to_excel(data_dict, template_file="template-2.xlsx", output_file="output.xlsx"):
    """
    1) Open 'template_file' and activate the sheet named 'Inputs'.
    2) Search all cells in 'Inputs' for a matching key from data_dict.
    3) If a cell's value matches a key, get the dictionary value and:
       - If it starts with "$" (optional spaces allowed), parse as float and apply currency format.
       - If it ends with "%", parse as float/100 and apply percentage format.
       - If it's strictly numeric (with optional commas/decimal), parse as float.
       - Otherwise, write as text.
    4) Save the modified workbook as 'output_file'.
    """
    wb = load_workbook(template_file)
    sheet = wb["Inputs"]
    max_row = sheet.max_row
    max_col = sheet.max_column

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell_value = cell.value

            if cell_value is not None:
                field_label = str(cell_value).strip()

                if field_label in data_dict:
                    out_cell = sheet.cell(row=row, column=col + 1)
                    raw_value = str(data_dict[field_label]).strip()

                    # 1) Currency format (allows space after $)
                    if dollar_pattern.match(raw_value):
                        # Remove '$', any spaces, and commas
                        numeric_str = raw_value.replace("$", "").replace(" ", "").replace(",", "")
                        numeric_val = float(numeric_str)
                        out_cell.value = numeric_val
                        out_cell.number_format = '$#,##0.00'
                        continue

                    # 2) Percentage format
                    if percent_pattern.match(raw_value):
                        numeric_str = raw_value.replace("%", "").replace(",", "")
                        numeric_val = float(numeric_str) / 100.0
                        out_cell.value = numeric_val
                        out_cell.number_format = '0.00%'
                        continue

                    # 3) Generic numeric
                    if numeric_pattern.match(raw_value):
                        numeric_str = raw_value.replace(",", "")
                        numeric_val = float(numeric_str)
                        out_cell.value = numeric_val
                        out_cell.number_format = 'General'
                        continue

                    # 4) Fallback to text
                    out_cell.value = raw_value

    wb.save(output_file)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({'error': 'No PDF path provided'}))
        sys.exit(1)

    PDF_PATH = sys.argv[1]
    
    # PDF_PATH = "rmbs_file_scrape.pdf"
    AZURE_API_KEY = os.getenv("API_KEY")
    AZURE_ENDPOINT = os.getenv("AZURE_ENDPOINT")
    DEPLOYMENT_ID = os.getenv("DEPLOYMENT_NAME")
    API_VERSION = os.getenv("API_VERSION")

    json_output = pdf_to_json(PDF_PATH, AZURE_API_KEY, AZURE_ENDPOINT, DEPLOYMENT_ID, API_VERSION)
    print(json.dumps(json_output, indent=4))
    save_to_excel(json_output)